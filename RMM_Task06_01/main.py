import os
import sys
from pathlib import Path

import openpyxl as openpyxl
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage
from openpyxl import load_workbook


class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('uis/main.ui', self)
        self.setWindowIcon(QtGui.QIcon('images/icon.png'))
        self.setWindowTitle('Создание копий визитки в Word и Excel с помощью Python')

        self.btn_wrd.clicked.connect(self.execute_wrd)
        self.btn_xl.clicked.connect(self.execute_xl)

    def execute_wrd(self):
        document_path = Path(__file__).parent / "temper.docx"

        doc = DocxTemplate(document_path)
        s = Path(__file__).parent / 'images' / "icon.png"
        filepath =  str(s)
        imagen = InlineImage(doc, filepath, width=Mm(5.0),height=Mm(5.0))
        context = {"logo": imagen,
                   "First_Last_Name": self.nameLine.text(),
                   "Employe": self.employeLine.text(),
                   "Company": self.companyLine.text(),
                   "Telega": self.telegramLine.text(),
                   "Insta": self.instagramLine.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "generated.docx")
        os.system('start generated.docx')

    def execute_xl(self):
        fn = 'temper.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['B3'] = self.nameLine.text()
        ws['B4'] = self.employeLine.text()
        ws['B6'] = self.companyLine.text()
        ws['B8'] = self.telegramLine.text()
        ws['B9'] = self.instagramLine.text()
        print(Path(__file__).parent / 'images' / "icon.png")
        img = openpyxl.drawing.image.Image(Path(__file__).parent / 'images' / "Free_Icon.jpg")
        img.anchor = 'C1'
        img.width = 40
        img.height = 40
        ws.add_image(img)
        wb.save(Path(__file__).parent / "generated.xlsx")
        wb.close()
        os.system('start generated.xlsx')

def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
